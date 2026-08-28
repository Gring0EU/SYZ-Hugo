import datetime, json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule

import os
OUT     = os.environ.get('OUT_NAME','SYZ_Spread_Since_14Jul2026.xlsx')
ANCHOR  = datetime.datetime(2026, 7, 14)
ANCHOR_TXT = '14.07.2026'
FIRST   = 7                      # first data row
LAST    = int(os.environ.get('LAST_ROW','850'))                    # formulas extend this far (blank rows stay blank)

d = json.load(open('data.json'))
static, extra, desc = d['static'], d['extra'], d['desc']
isins = ([str(r[0]).strip() for r in static] + [str(e).strip() for e in extra])[:int(os.environ.get('N_ISIN','100000'))]   # source list carries stray spaces; BDP needs them gone

FONT   = 'Arial'
INK    = 'FF1F2933'
ACCENT = 'FF1F4E79'
def F(sz=10, b=False, i=False, color=INK): return Font(name=FONT, sz=sz, b=b, i=i, color=color)
THIN   = Side(style='thin', color='FFD0D7DE')
BOX    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = openpyxl.Workbook()

# =============================================================== Spread Monitor
ws = wb.active
ws.title = 'Spread Monitor'
ws.sheet_properties.tabColor = ACCENT

ws['A1'] = 'Bond spread variation since ' + ANCHOR_TXT
ws['A1'].font = F(16, b=True, color=ACCENT)
ws['A2'] = 'Anchor date'
ws['A2'].font = F(10, b=True)
ws['B2'] = ANCHOR
ws['B2'].number_format = 'DD.MM.YYYY'
ws['B2'].font = F(10, b=True, color='FF0000FF')
ws['B2'].fill = PatternFill('solid', fgColor='FFFFFF00')
ws['B2'].border = BOX
ws['B2'].comment = Comment(
    'The only input cell on this sheet. Every "since" column reads it as the Bloomberg\n'
    'SETTLE_DT. Change it to re-base the whole analysis - the headers in row 6 follow\n'
    'automatically. Requested value: 14 July 2026.', 'Claude')
ws['C2'] = 'Blue on yellow = the only cell to edit. Everything else is a formula.'
ws['C2'].font = F(9, i=True, color='FF6B7280')
ws['A3'] = 'Bonds priced'
ws['A3'].font = F(10, b=True)
ws['B3'] = '=COUNT($S${f}:$S${l})&" of "&COUNTA($A${f}:$A${l})'.format(f=FIRST, l=LAST)
ws['B3'].font = F(10)
ws['C3'] = 'Paste additional ISINs at the bottom of column A - formulas already run to row {}.'.format(LAST)
ws['C3'].font = F(9, i=True, color='FF6B7280')
ws['A4'] = 'Data'
ws['A4'].font = F(10, b=True)
ws['B4'] = 'Bloomberg BDP (add-in must be connected). Press F9 / Ctrl+Alt+F9 to refresh.'
ws['B4'].font = F(9, i=True, color='FF6B7280')

# column letter -> (header, number format, width, formula builder)
BDP = '_xll.BDP($A{r}&" Corp",{a})'
def g(r, body): return '=IF($A{r}="","",{b})'.format(r=r, b=body)
ST_END = len(static) + 60        # SYZ_static data extent, plus room to grow
CD_END = len(desc) + 60          # Company_desc data extent, plus room to grow
LK  = 'IFERROR(INDEX(SYZ_static!${c}$2:${c}$%d,MATCH($A{r},SYZ_static!$A$2:$A$%d,0)),"")' % (ST_END, ST_END)

SPEC = [
 ('A', 'ISIN',                       '@',          16, None),
 ('B', 'Company',                    '@',          30, lambda r: g(r, 'PROPER(' + BDP.format(r=r, a='"NAME"') + ')')),
 ('C', 'Company description',        '@',          58, lambda r: g(r, 'IFERROR(INDEX(Company_desc!$B$2:$B$%d,MATCH($A{r},Company_desc!$A$2:$A$%d,0)),{lk})'.format(r=r, lk=LK.format(c='Q', r=r)) % (CD_END, CD_END))),
 ('D', 'SYZ view',                   '@',          14, lambda r: g(r, LK.format(c='B', r=r))),
 ('E', 'Sector',                     '@',          20, lambda r: g(r, BDP.format(r=r, a='"INDUSTRY_SECTOR"'))),
 ('F', 'Country',                    '@',          16, lambda r: g(r, 'PROPER(' + BDP.format(r=r, a='"COUNTRY_FULL_NAME"') + ')')),
 ('G', 'Ccy',                        '@',           7, lambda r: g(r, BDP.format(r=r, a='"CRNCY"'))),
 ('H', 'Rating',                     '@',          10, lambda r: g(r, 'IFERROR(INDEX(SYZ_static!$T$2:$T$%d,MATCH($A{r},SYZ_static!$A$2:$A$%d,0)),{b})'.format(r=r, b=BDP.format(r=r, a='"RTG_SP"')) % (ST_END, ST_END))),
 ('I', 'Classification',             '@',          18, lambda r: g(r, LK.format(c='S', r=r))),
 ('J', 'Structure',                  '@',          18, lambda r: g(r, BDP.format(r=r, a='"PAYMENT_RANK"'))),
 ('K', 'Coupon',                     '0.000',      10, lambda r: g(r, BDP.format(r=r, a='"CPN"'))),
 ('L', 'Maturity',                   'DD.MM.YYYY', 12, lambda r: g(r, BDP.format(r=r, a='"MATURITY"'))),
 ('M', 'Den. (k)',                   '0.###',      10, lambda r: g(r, 'IFERROR(' + BDP.format(r=r, a='"MIN_PIECE"') + '/1000,"")')),
 ('N', 'Price',                      '0.00',       10, lambda r: g(r, BDP.format(r=r, a='"PX_LAST"'))),
 ('O', 'YTW (%)',                    '0.00',       10, lambda r: g(r, BDP.format(r=r, a='"YAS_BOND_YLD","YAS_YLD_FLAG=15"'))),
 ('P', 'Mod. dur.',                  '0.00',       10, lambda r: g(r, BDP.format(r=r, a='"YAS_MOD_DUR"'))),
 ('Q', 'Spread today (bp)',          '0',          15, lambda r: g(r, BDP.format(r=r, a='"YAS_OAS_SPRD"'))),
 ('R', 'Spread on anchor (bp)',      '0',          17, lambda r: g(r, BDP.format(r=r, a='"YAS_OAS_SPRD","SETTLE_DT",$B$2'))),
 ('S', 'Spread change (bp)',         '+0;-0;0',    16, lambda r: g(r, 'IF(OR(NOT(ISNUMBER($Q{r})),NOT(ISNUMBER($R{r}))),"",$Q{r}-$R{r})'.format(r=r))),
 ('T', 'YTW on anchor (%)',          '0.00',       15, lambda r: g(r, BDP.format(r=r, a='"YAS_BOND_YLD","YAS_YLD_FLAG=15","SETTLE_DT",$B$2'))),
 ('U', 'YTW change (%)',             '+0.00;-0.00;0.00', 14, lambda r: g(r, 'IF(OR(NOT(ISNUMBER($O{r})),NOT(ISNUMBER($T{r}))),"",$O{r}-$T{r})'.format(r=r))),
 ('V', 'Rank widening',              '0',          13, lambda r: '=IF(NOT(ISNUMBER($S{r})),"",COUNTIFS($G${f}:$G${l},$G{r},$S${f}:$S${l},">"&$S{r})+COUNTIFS($G${f}:$G{r},$G{r},$S${f}:$S{r},$S{r}))'.format(r=r, f=FIRST, l=LAST)),
 ('W', 'Rank tightening',            '0',          13, lambda r: '=IF(NOT(ISNUMBER($S{r})),"",COUNTIFS($G${f}:$G${l},$G{r},$S${f}:$S${l},"<"&$S{r})+COUNTIFS($G${f}:$G{r},$G{r},$S${f}:$S{r},$S{r}))'.format(r=r, f=FIRST, l=LAST)),
 ('X', 'Key widening',               '@',          14, lambda r: '=IF($V{r}="","",$G{r}&"|"&$V{r})'.format(r=r)),
 ('Y', 'Key tightening',             '@',          14, lambda r: '=IF($W{r}="","",$G{r}&"|"&$W{r})'.format(r=r)),
]
DYNAMIC_HDR = {'R': '="Spread on "&TEXT($B$2,"dd.mm.yyyy")&" (bp)"',
               'S': '="Spread change since "&TEXT($B$2,"dd.mm.yyyy")&" (bp)"',
               'T': '="YTW on "&TEXT($B$2,"dd.mm.yyyy")&" (%)"',
               'U': '="YTW change since "&TEXT($B$2,"dd.mm.yyyy")&" (%)"'}

hdr_fill = PatternFill('solid', fgColor=ACCENT)
for letter, header, fmt, width, _ in SPEC:
    c = ws[letter + '6']
    c.value = DYNAMIC_HDR.get(letter, header)
    c.font = F(10, b=True, color='FFFFFFFF')
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = BOX
    ws.column_dimensions[letter].width = width
ws.row_dimensions[6].height = 32

alt = PatternFill('solid', fgColor='FFF6F8FA')
for i, r in enumerate(range(FIRST, LAST + 1)):
    isin = isins[i] if i < len(isins) else None
    for letter, header, fmt, width, fn in SPEC:
        c = ws[letter + str(r)]
        c.value = isin if letter == 'A' else (fn(r) if fn else None)
        c.font = F(9, color='FF0000FF' if letter == 'A' else INK)
        c.number_format = fmt
        c.border = BOX
        c.alignment = Alignment(horizontal='left' if fmt == '@' else 'right', vertical='center')
        if r % 2 == 0:
            c.fill = alt
ws.freeze_panes = 'C7'
ws.auto_filter.ref = 'A6:Y{}'.format(LAST)
ws.column_dimensions.group('V', 'Y', hidden=True)      # ranking helpers, collapsed
ws.sheet_view.showGridLines = False

# red widening / green tightening on the two change columns
for rng in ('S{}:S{}'.format(FIRST, LAST), 'U{}:U{}'.format(FIRST, LAST)):
    ws.conditional_formatting.add(rng, CellIsRule(operator='greaterThan', formula=['0'],
                                                  font=Font(name=FONT, sz=9, color='FFB42318')))
    ws.conditional_formatting.add(rng, CellIsRule(operator='lessThan', formula=['0'],
                                                  font=Font(name=FONT, sz=9, color='FF067647')))

# =============================================================== Top & Bottom
tb = wb.create_sheet('Top & Bottom')
tb.sheet_properties.tabColor = 'FF2E7D32'
tb.sheet_view.showGridLines = False
tb['A1'] = 'Biggest spread movers since ' + ANCHOR_TXT
tb['A1'].font = F(16, b=True, color=ACCENT)
tb['A2'] = '=\'Spread Monitor\'!$B$2'
tb['A2'].number_format = '"Anchor date: "DD.MM.YYYY'
tb['A2'].font = F(10, i=True, color='FF6B7280')
tb['A3'] = 'Ranked within each currency. Live formulas - they follow the Spread Monitor sheet automatically.'
tb['A3'].font = F(9, i=True, color='FF6B7280')

# display column -> source column on Spread Monitor
DISP = [('ISIN', 'A', '@', 16), ('Company', 'B', '@', 28), ('Company description', 'C', '@', 52),
        ('Sector', 'E', '@', 18), ('Ccy', 'G', '@', 7), ('Rating', 'H', '@', 9),
        ('Maturity', 'L', 'DD.MM.YYYY', 12), ('YTW (%)', 'O', '0.00', 9),
        ('Spread today (bp)', 'Q', '0', 14), ('Spread on anchor (bp)', 'R', '0', 16),
        ('Spread change (bp)', 'S', '+0;-0;0', 15)]
for j, (h, src, fmt, width) in enumerate(DISP, start=1):
    tb.column_dimensions[get_column_letter(j)].width = width

TOPN = 10
row = 5
for ccy in ('USD', 'EUR', 'GBP', 'CHF'):
    for label, keycol, tone in (('widened the most', 'X', 'FFB42318'), ('tightened the most', 'Y', 'FF067647')):
        tb.cell(row, 1, '{c} - Top {n} that {l} since {d}'.format(c=ccy, n=TOPN, l=label, d=ANCHOR_TXT))
        tb.cell(row, 1).font = F(12, b=True, color=tone)
        row += 1
        for j, (h, src, fmt, width) in enumerate(DISP, start=1):
            c = tb.cell(row, j, h)
            c.font = F(9, b=True, color='FFFFFFFF')
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = BOX
        tb.row_dimensions[row].height = 26
        row += 1
        for i in range(1, TOPN + 1):
            for j, (h, src, fmt, width) in enumerate(DISP, start=1):
                c = tb.cell(row, j)
                c.value = ('=IFERROR(INDEX(\'Spread Monitor\'!${s}${f}:${s}${l},'
                           'MATCH("{ccy}|{i}",\'Spread Monitor\'!${k}${f}:${k}${l},0)),"")'
                           ).format(s=src, f=FIRST, l=LAST, ccy=ccy, i=i, k=keycol)
                c.font = F(9)
                c.number_format = fmt
                c.border = BOX
                c.alignment = Alignment(horizontal='left' if fmt == '@' else 'right', vertical='center')
                if i % 2 == 0:
                    c.fill = alt
            row += 1
        row += 2

# =============================================================== input data sheets
cds = wb.create_sheet('Company_desc')
cds.sheet_properties.tabColor = 'FF9AA5B1'
for j, h in enumerate(['ISIN', 'Company description'], start=1):
    c = cds.cell(1, j, h); c.font = F(10, b=True, color='FFFFFFFF'); c.fill = hdr_fill; c.border = BOX
for i, (a, b) in enumerate(desc, start=2):
    cds.cell(i, 1, str(a).strip()).font = F(9)
    cds.cell(i, 2, b).font = F(9)
cds.column_dimensions['A'].width = 16
cds.column_dimensions['B'].width = 90
cds.freeze_panes = 'A2'

st = wb.create_sheet('SYZ_static')
st.sheet_properties.tabColor = 'FF9AA5B1'
HEAD = ['ISIN','SYZ view','Company','Coupon','Structure','Maturity','Country','Sector','Den. (k)',
        'Price','YTW','YTW 1w chg','YTW 1m chg','Spread','Dur.','Ind LV.','Company Description',
        'Currency','Bond Classification','rating','Tax Group','Column1']
for j, h in enumerate(HEAD, start=1):
    c = st.cell(1, j, h); c.font = F(10, b=True, color='FFFFFFFF'); c.fill = hdr_fill; c.border = BOX
for i, r in enumerate(static, start=2):
    for j, v in enumerate(r, start=1):
        st.cell(i, j, str(v).strip() if j == 1 and v is not None else v).font = F(9)
st.column_dimensions['A'].width = 16
st.freeze_panes = 'A2'

# =============================================================== ReadMe
rm = wb.create_sheet('ReadMe')
rm.sheet_properties.tabColor = 'FFF59E0B'
rm.sheet_view.showGridLines = False
lines = [
 ('Bond spread variation since ' + ANCHOR_TXT, ''),
 ('', ''),
 ('SHEETS', ''),
 ('Spread Monitor', 'One row per bond ({n} ISINs). Every column except A is a formula. Filter and sort with the '
                    'arrows on row 6 - no macro involved.'.format(n=len(isins))),
 ('Top & Bottom', 'Top 10 wideners and Top 10 tighteners per currency (USD, EUR, GBP, CHF), pulled live from '
                  'Spread Monitor with INDEX/MATCH on the hidden ranking keys (columns V:Y).'),
 ('Company_desc', 'Input data: the ISIN -> company-description library carried over from the original workbook.'),
 ('SYZ_static', 'Input data: the SYZ-internal fields (SYZ view, Ind LV., Bond Classification, Tax Group, rating) '
                'that are not available on Bloomberg, carried over from the original workbook.'),
 ('', ''),
 ('HOW TO USE', ''),
 ('1', 'Open with the Bloomberg add-in connected, then Ctrl+Alt+F9 to force a full recalculation.'),
 ('2', 'Change the anchor date in Spread Monitor!B2 (blue on yellow) to re-base everything. '
       'The four "since" headers rewrite themselves.'),
 ('3', 'Add bonds by pasting ISINs at the bottom of column A - all formulas already run to row {}.'.format(LAST)),
 ('', ''),
 ('BLOOMBERG FIELDS', ''),
 ('Company', 'PROPER(BDP(ISIN & " Corp","NAME"))'),
 ('Sector', 'BDP("INDUSTRY_SECTOR")'),
 ('Country', 'PROPER(BDP("COUNTRY_FULL_NAME"))'),
 ('Ccy', 'BDP("CRNCY")'),
 ('Structure', 'BDP("PAYMENT_RANK")'),
 ('Coupon / Maturity', 'BDP("CPN") / BDP("MATURITY")'),
 ('Den. (k)', 'BDP("MIN_PIECE")/1000'),
 ('Price', 'BDP("PX_LAST")'),
 ('YTW (%)', 'BDP("YAS_BOND_YLD","YAS_YLD_FLAG=15")'),
 ('Mod. dur.', 'BDP("YAS_MOD_DUR")'),
 ('Spread today (bp)', 'BDP("YAS_OAS_SPRD")'),
 ('Spread on anchor (bp)', 'BDP("YAS_OAS_SPRD","SETTLE_DT",$B$2)'),
 ('YTW on anchor (%)', 'BDP("YAS_BOND_YLD","YAS_YLD_FLAG=15","SETTLE_DT",$B$2)'),
 ('Spread change / YTW change', 'Plain subtraction of the two columns above - not a third Bloomberg call.'),
 ('Rating', 'SYZ_static lookup, falling back to BDP("RTG_SP") for ISINs not in that sheet.'),
 ('Company description / SYZ view / Classification', 'INDEX/MATCH on Company_desc and SYZ_static by ISIN.'),
 ('', ''),
 ('NOTES', ''),
 ('Not recalculated offline', 'Every value comes from the Bloomberg add-in (_xll.BDP), which resolves only on a '
                              'terminal-connected machine. Cells read as 0 or blank until you refresh in Excel.'),
 ('Blank-guarded', 'Every formula is wrapped in IF($A="","",...), so the unused rows below the last ISIN stay empty '
                   'instead of filling the sheet with errors.'),
 ('Ranking columns', 'V:Y on Spread Monitor are grouped and hidden. Rank = COUNTIFS within currency, with a '
                     'running COUNTIFS tie-break so two bonds with the same move never share a rank.'),
 ('ISIN universe', 'The 746 bonds from the original Sheet1 plus the 17 newer ISINs that were sitting on the '
                   'original Formulas sheet = {n} in total.'.format(n=len(isins))),
]
rm['A1'].font = F(16, b=True, color=ACCENT)
for i, (a, b) in enumerate(lines, start=1):
    rm.cell(i, 1, a); rm.cell(i, 2, b)
    if a and not b:
        rm.cell(i, 1).font = F(12, b=True, color=ACCENT)
    else:
        rm.cell(i, 1).font = F(10, b=True)
        rm.cell(i, 2).font = F(10)
    rm.cell(i, 2).alignment = Alignment(wrap_text=True, vertical='top')
rm['A1'].font = F(16, b=True, color=ACCENT)
rm.column_dimensions['A'].width = 34
rm.column_dimensions['B'].width = 105

wb.save(OUT)
print('saved', OUT, '| bonds:', len(isins))
