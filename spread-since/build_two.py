import datetime, json, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

OUT     = os.environ.get('OUT_NAME', 'Spread_Variation_USD_EUR_since_14Jul2026.xlsx')
ANCHOR  = datetime.datetime(2026, 7, 14)
ATXT    = '14.07.2026'
SPARE   = 6                      # blank-but-live rows at the end of each block

lists = json.load(open('lists.json'))

FONT, INK, ACCENT = 'Arial', 'FF1F2933', 'FF1F4E79'
def F(sz=10, b=False, i=False, color=INK): return Font(name=FONT, sz=sz, b=b, i=i, color=color)
THIN = Side(style='thin', color='FFD0D7DE')
BOX  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR  = PatternFill('solid', fgColor=ACCENT)
ALT  = PatternFill('solid', fgColor='FFF6F8FA')

BDP = '_xll.BDP($A{r}&" Corp",{a})'
def g(r, body): return '=IF($A{r}="","",{b})'.format(r=r, b=body)

# letter, header, number format, width, formula builder
SPEC = [
 ('A', 'ISIN',                  '@',           15, None),
 ('B', 'Security',              '@',           26, lambda r: g(r, BDP.format(r=r, a='"SECURITY_DES"'))),
 ('C', 'Company',               '@',           28, lambda r: g(r, 'PROPER(' + BDP.format(r=r, a='"NAME"') + ')')),
 ('D', 'Company description',   '@',           60, lambda r: g(r, 'IFERROR(' + BDP.format(r=r, a='"CIE_DES"') + ',"")')),
 ('E', 'Sector',                '@',           20, lambda r: g(r, BDP.format(r=r, a='"INDUSTRY_SECTOR"'))),
 ('F', 'Industry group',        '@',           22, lambda r: g(r, BDP.format(r=r, a='"INDUSTRY_GROUP"'))),
 ('G', 'Country',               '@',           16, lambda r: g(r, 'PROPER(' + BDP.format(r=r, a='"COUNTRY_FULL_NAME"') + ')')),
 ('H', 'Ccy',                   '@',            6, lambda r: g(r, BDP.format(r=r, a='"CRNCY"'))),
 ('I', 'Rating (comp.)',        '@',           12, lambda r: g(r, 'IFERROR(' + BDP.format(r=r, a='"BB_COMPOSITE"') + ',' + BDP.format(r=r, a='"RTG_SP"') + ')')),
 ('J', 'Structure',             '@',           18, lambda r: g(r, BDP.format(r=r, a='"PAYMENT_RANK"'))),
 ('K', 'Coupon',                '0.000',       9,  lambda r: g(r, BDP.format(r=r, a='"CPN"'))),
 ('L', 'Maturity',              'DD.MM.YYYY',  12, lambda r: g(r, BDP.format(r=r, a='"MATURITY"'))),
 ('M', 'Den. (k)',              '0.###',       9,  lambda r: g(r, 'IFERROR(' + BDP.format(r=r, a='"MIN_PIECE"') + '/1000,"")')),
 ('N', 'Price',                 '0.00',        9,  lambda r: g(r, BDP.format(r=r, a='"PX_LAST"'))),
 ('O', 'YTW (%)',               '0.00',        9,  lambda r: g(r, BDP.format(r=r, a='"YAS_BOND_YLD","YAS_YLD_FLAG=15"'))),
 ('P', 'Mod. dur.',             '0.00',        9,  lambda r: g(r, BDP.format(r=r, a='"YAS_MOD_DUR"'))),
 ('Q', 'Spread today (bp)',     '0',           14, lambda r: g(r, BDP.format(r=r, a='"YAS_OAS_SPRD"'))),
 ('R', 'Spread @ anchor (bp)',  '0',           15, lambda r: g(r, BDP.format(r=r, a='"YAS_OAS_SPRD","SETTLE_DT",$B$2'))),
 ('S', 'Spread change (bp)',    '+0;-0;0',     15, lambda r: g(r, 'IF(OR(NOT(ISNUMBER($Q{r})),NOT(ISNUMBER($R{r}))),"",$Q{r}-$R{r})'.format(r=r))),
 ('T', 'YTW @ anchor (%)',      '0.00',        14, lambda r: g(r, BDP.format(r=r, a='"YAS_BOND_YLD","YAS_YLD_FLAG=15","SETTLE_DT",$B$2'))),
 ('U', 'YTW change (%)',        '+0.00;-0.00;0.00', 13, lambda r: g(r, 'IF(OR(NOT(ISNUMBER($O{r})),NOT(ISNUMBER($T{r}))),"",$O{r}-$T{r})'.format(r=r))),
]
LASTCOL = SPEC[-1][0]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Spread since ' + ATXT
ws.sheet_properties.tabColor = ACCENT
ws.sheet_view.showGridLines = False

ws['A1'] = 'Spread variation since ' + ATXT + ' - USD and EUR lists'
ws['A1'].font = F(16, b=True, color=ACCENT)
ws['A2'] = 'Anchor date'
ws['A2'].font = F(10, b=True)
ws['B2'] = ANCHOR
ws['B2'].number_format = 'DD.MM.YYYY'
ws['B2'].font = F(10, b=True, color='FF0000FF')
ws['B2'].fill = PatternFill('solid', fgColor='FFFFFF00')
ws['B2'].border = BOX
ws['B2'].comment = Comment(
    'The only input cell. Both tables pass it to Bloomberg as SETTLE_DT, and the four\n'
    '"since" headers derive their labels from it, so changing this date re-bases\n'
    'everything. Requested value: 14 July 2026.', 'Claude')
ws['C2'] = 'Blue on yellow = the only cell to edit. Every other cell is a formula.'
ws['C2'].font = F(9, i=True, color='FF6B7280')
ws['A3'] = 'Data'
ws['A3'].font = F(10, b=True)
ws['B3'] = 'Bloomberg BDP - open with the add-in connected, then Ctrl+Alt+F9 to recalculate.'
ws['B3'].font = F(9, i=True, color='FF6B7280')

for letter, header, fmt, width, fn in SPEC:
    ws.column_dimensions[letter].width = width

def block(title, isins, start, tblname, tone):
    """Write one section: title, header row, data rows (+ spare), summary."""
    ws.cell(start, 1, '="{t}  -  spread variation since "&TEXT($B$2,"dd.mm.yyyy")'.format(t=title))
    ws.cell(start, 1).font = F(13, b=True, color=tone)
    hrow = start + 1
    for letter, header, fmt, width, fn in SPEC:
        c = ws[letter + str(hrow)]
        c.value = header
        c.font = F(9, b=True, color='FFFFFFFF')
        c.fill = HDR
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BOX
    ws.row_dimensions[hrow].height = 30

    first = hrow + 1
    last = first + len(isins) + SPARE - 1
    for n, r in enumerate(range(first, last + 1)):
        isin = isins[n] if n < len(isins) else None
        for letter, header, fmt, width, fn in SPEC:
            c = ws[letter + str(r)]
            c.value = isin if letter == 'A' else fn(r)
            c.font = F(9, color='FF0000FF' if letter == 'A' else INK)
            c.number_format = fmt
            c.border = BOX
            c.alignment = Alignment(horizontal='left' if fmt == '@' else 'right',
                                    vertical='center', wrap_text=(letter == 'D'))
            if n % 2 == 1:
                c.fill = ALT

    tbl = Table(displayName=tblname, ref='A{}:{}{}'.format(hrow, LASTCOL, last))
    tbl.tableStyleInfo = TableStyleInfo(name='TableStyleLight1', showRowStripes=False)
    ws.add_table(tbl)

    for rng in ('S{}:S{}'.format(first, last), 'U{}:U{}'.format(first, last)):
        ws.conditional_formatting.add(rng, CellIsRule(operator='greaterThan', formula=['0'],
                                     font=Font(name=FONT, sz=9, b=True, color='FFB42318')))
        ws.conditional_formatting.add(rng, CellIsRule(operator='lessThan', formula=['0'],
                                     font=Font(name=FONT, sz=9, b=True, color='FF067647')))

    s = start + len(isins) + SPARE + 3          # summary block, two rows below the table
    D = '$S${f}:$S${l}'.format(f=first, l=last)
    A = '$A${f}:$A${l}'.format(f=first, l=last)
    C = '$C${f}:$C${l}'.format(f=first, l=last)
    rows = [
      ('Bonds priced',            '=COUNT({d})&" of "&COUNTA($A${f}:$A${l})'.format(d=D, f=first, l=last), '@'),
      ('Average spread change',   '=IFERROR(AVERAGE({d}),"")'.format(d=D), '+0.0;-0.0;0.0'),
      ('Median spread change',    '=IFERROR(MEDIAN({d}),"")'.format(d=D), '+0.0;-0.0;0.0'),
      ('Widened / tightened',     '=COUNTIF({d},">0")&" wider / "&COUNTIF({d},"<0")&" tighter"'.format(d=D), '@'),
      ('Biggest widening',        '=IFERROR(MAX({d}),"")'.format(d=D), '+0;-0;0'),
      ('  ... which bond',        '=IFERROR(INDEX({a},MATCH(MAX({d}),{d},0))&" - "&INDEX({c},MATCH(MAX({d}),{d},0)),"")'.format(a=A, c=C, d=D), '@'),
      ('Biggest tightening',      '=IFERROR(MIN({d}),"")'.format(d=D), '+0;-0;0'),
      ('  ... which bond',        '=IFERROR(INDEX({a},MATCH(MIN({d}),{d},0))&" - "&INDEX({c},MATCH(MIN({d}),{d},0)),"")'.format(a=A, c=C, d=D), '@'),
    ]
    ws.cell(s - 1, 1, title.split(' ')[0] + ' summary')
    ws.cell(s - 1, 1).font = F(11, b=True, color=tone)
    for i, (label, formula, fmt) in enumerate(rows):
        lc = ws.cell(s + i, 1, label); lc.font = F(9, b=True); lc.border = BOX
        vc = ws.cell(s + i, 2, formula); vc.font = F(9); vc.number_format = fmt; vc.border = BOX
        vc.alignment = Alignment(horizontal='left' if fmt == '@' else 'right')
        ws.merge_cells(start_row=s + i, start_column=2, end_row=s + i, end_column=4)
    return s + len(rows) + 3

nxt = block('USD bonds ({} ISINs)'.format(len(lists['USD'])), lists['USD'], 5, 'tblUSD', 'FF1F4E79')
block('EUR bonds ({} ISINs)'.format(len(lists['EUR'])), lists['EUR'], nxt, 'tblEUR', 'FF6B21A8')

ws.freeze_panes = 'B6'

# ---------------------------------------------------------------- ReadMe
rm = wb.create_sheet('ReadMe')
rm.sheet_properties.tabColor = 'FFF59E0B'
rm.sheet_view.showGridLines = False
lines = [
 ('Spread variation since ' + ATXT, ''),
 ('', ''),
 ('WHAT THIS IS', ''),
 ('Scope', 'The two lists you supplied: {u} USD bonds and {e} EUR bonds. Nothing else.'.format(
            u=len(lists['USD']), e=len(lists['EUR']))),
 ('One sheet, two tables', 'tblUSD and tblEUR, each a real Excel table with its own filter arrows, '
                           'so you can sort either block by spread change without touching the other.'),
 ('Anchor date', 'Cell B2 (blue on yellow) is the only input. Both tables pass it to Bloomberg as '
                 'SETTLE_DT, and the two section titles restate it. The four "since" columns are '
                 'headed "@ anchor" rather than a spelled-out date, because an Excel table header '
                 'cannot hold a formula - so nothing goes stale when you change B2.'),
 ('Spare rows', '{} live but empty rows at the bottom of each table - paste an ISIN in column A and '
                'the row fills itself.'.format(SPARE)),
 ('', ''),
 ('BLOOMBERG FIELDS', ''),
 ('B  Security', 'BDP(ISIN & " Corp","SECURITY_DES")'),
 ('C  Company', 'PROPER(BDP("NAME"))'),
 ('D  Company description', 'BDP("CIE_DES")'),
 ('E  Sector / F Industry group', 'BDP("INDUSTRY_SECTOR") / BDP("INDUSTRY_GROUP")'),
 ('G  Country', 'PROPER(BDP("COUNTRY_FULL_NAME"))'),
 ('H  Ccy', 'BDP("CRNCY")'),
 ('I  Rating (comp.)', 'BDP("BB_COMPOSITE"), falling back to BDP("RTG_SP")'),
 ('J  Structure', 'BDP("PAYMENT_RANK")'),
 ('K  Coupon / L Maturity', 'BDP("CPN") / BDP("MATURITY")'),
 ('M  Den. (k)', 'BDP("MIN_PIECE")/1000'),
 ('N  Price', 'BDP("PX_LAST")'),
 ('O  YTW (%)', 'BDP("YAS_BOND_YLD","YAS_YLD_FLAG=15")'),
 ('P  Mod. dur.', 'BDP("YAS_MOD_DUR")'),
 ('Q  Spread today (bp)', 'BDP("YAS_OAS_SPRD")'),
 ('R  Spread on anchor (bp)', 'BDP("YAS_OAS_SPRD","SETTLE_DT",$B$2)'),
 ('S  Spread change (bp)', 'Q - R. Positive = wider (red), negative = tighter (green).'),
 ('T  YTW on anchor (%)', 'BDP("YAS_BOND_YLD","YAS_YLD_FLAG=15","SETTLE_DT",$B$2)'),
 ('U  YTW change (%)', 'O - T'),
 ('', ''),
 ('NOTES', ''),
 ('Refresh', 'Open with the Bloomberg add-in connected and press Ctrl+Alt+F9. Until then the cells '
             'are blank - _xll.BDP only resolves on a terminal.'),
 ('Blank-guarded', 'Every formula is wrapped in IF($A="","",...), so the spare rows stay empty rather '
                   'than showing errors.'),
 ('Changes are subtractions', 'Spread change and YTW change subtract the two visible columns instead of '
                              'making a third Bloomberg call, so the three numbers on a row always agree.'),
 ('If a field comes back #N/A', 'CIE_DES and BB_COMPOSITE are not populated for every issuer; those two '
                                'are wrapped in IFERROR so one missing field cannot blank the row.'),
]
rm['A1'].font = F(16, b=True, color=ACCENT)
for i, (a, b) in enumerate(lines, start=1):
    rm.cell(i, 1, a); rm.cell(i, 2, b)
    if a and not b:
        rm.cell(i, 1).font = F(12, b=True, color=ACCENT)
    else:
        rm.cell(i, 1).font = F(10, b=True); rm.cell(i, 2).font = F(10)
    rm.cell(i, 2).alignment = Alignment(wrap_text=True, vertical='top')
rm['A1'].font = F(16, b=True, color=ACCENT)
rm.column_dimensions['A'].width = 30
rm.column_dimensions['B'].width = 100

wb.save(OUT)
print('saved', OUT)
