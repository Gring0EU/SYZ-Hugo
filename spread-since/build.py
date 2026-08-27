import datetime
from copy import copy
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

SRC = 'USD_orig.xlsm'
OUT = 'SYZ_Bond_Monitor_USD_spread_since_14Jul2026.xlsm'
ANCHOR = datetime.datetime(2026, 7, 14)
ANCHOR_TXT = '14.07.2026'
NEW_SHEET = 'Spread - since ' + ANCHOR_TXT           # 25 chars, <= 31
LAST_ROW = 769                                        # same extent as existing W/X formulas

wb = openpyxl.load_workbook(SRC, keep_vba=True)
f  = wb['Formulas']
blh = wb['Bonds_list_hard']
cd = wb['Company_desc']

# ---------------------------------------------------------------- 1. anchor date
cd['C1'] = 'FIXED ANCHOR DATE ->'
cd['C1'].font = Font(name='Nunito Sans', sz=9, b=True)
cd['D1'] = ANCHOR
cd['D1'].number_format = 'DD.MM.YYYY'
cd['D1'].font = Font(name='Nunito Sans', sz=9, b=True, color='FF0000FF')
cd['D1'].fill = PatternFill('solid', fgColor='FFFFFF00')
cd['C2'] = 'Edit D1 only. A1/A2 stay dynamic (TODAY()-7 / TODAY()-30).'
cd['C2'].font = Font(name='Nunito Sans', sz=9, i=True)
cd['D1'].comment = Comment(
    'Reference date for the "since" spread/yield columns (Formulas!Y:AA).\n'
    'Hardcoded input requested by the user: 14 July 2026.\n'
    'Change this cell to re-base the whole "since" analysis, then update the\n'
    'headers in Formulas!Y1:AA1 so the table labels stay in sync.', 'Claude')
cd.column_dimensions['C'].width = 24
cd.column_dimensions['D'].width = 14

# ---------------------------------------------------------------- 2. Formulas sheet
BDP = '_xll.BDP($A{r}&" Corp",{args})'
def g(r, body):                      # blank-guard so unused rows stay empty
    return '=IF($A{r}="","",{b})'.format(r=r, b=body)

def col_formula(letter, r):
    if letter == 'B':  return g(r, 'IFERROR(INDEX(Sheet1!$B:$B,MATCH($A{r},Sheet1!$A:$A,0)),"")'.format(r=r))
    if letter == 'C':  return g(r, 'PROPER(' + BDP.format(r=r, args='"NAME"') + ')')
    if letter == 'D':  return g(r, BDP.format(r=r, args='"CPN"'))
    if letter == 'E':  return g(r, BDP.format(r=r, args='"PAYMENT_RANK"'))
    if letter == 'F':  return g(r, BDP.format(r=r, args='"MATURITY"'))
    if letter == 'G':  return g(r, 'PROPER(' + BDP.format(r=r, args='"COUNTRY_FULL_NAME"') + ')')
    if letter == 'H':  return g(r, BDP.format(r=r, args='"INDUSTRY_SECTOR"'))
    if letter == 'I':  return g(r, 'IFERROR(' + BDP.format(r=r, args='"MIN_PIECE"') + '/1000,"")')
    if letter == 'J':  return g(r, BDP.format(r=r, args='"PX_LAST"'))
    if letter == 'K':  return g(r, BDP.format(r=r, args='"YAS_BOND_YLD","YAS_YLD_FLAG=15"'))
    if letter == 'L':  return g(r, 'K{r}-'.format(r=r) + BDP.format(r=r, args='"YAS_BOND_YLD","YAS_YLD_FLAG=15","SETTLE_DT",Company_desc!$A$1'))
    if letter == 'M':  return g(r, 'K{r}-'.format(r=r) + BDP.format(r=r, args='"YAS_BOND_YLD","YAS_YLD_FLAG=15","SETTLE_DT",Company_desc!$A$2'))
    if letter == 'N':  return g(r, BDP.format(r=r, args='"YAS_OAS_SPRD"'))
    if letter == 'O':  return g(r, BDP.format(r=r, args='"YAS_MOD_DUR"'))
    if letter == 'P':  return g(r, 'IFERROR(INDEX(Sheet1!$P:$P,MATCH($A{r},Sheet1!$A:$A,0)),"")'.format(r=r))
    if letter == 'Q':  return g(r, 'IFERROR(INDEX(Company_desc!$C:$C,MATCH($A{r},Company_desc!$A:$A,0)),'
                                   'IFERROR(INDEX(Sheet1!$Q:$Q,MATCH($A{r},Sheet1!$A:$A,0)),""))'.format(r=r))
    if letter == 'R':  return g(r, BDP.format(r=r, args='"CRNCY"'))
    if letter == 'S':  return g(r, 'IFERROR(INDEX(Sheet1!$S:$S,MATCH($A{r},Sheet1!$A:$A,0)),"")'.format(r=r))
    if letter == 'T':  return g(r, 'IFERROR(INDEX(Sheet1!$T:$T,MATCH($A{r},Sheet1!$A:$A,0)),'.format(r=r)
                                   + BDP.format(r=r, args='"RTG_SP"') + ')')
    if letter == 'U':  return g(r, 'IFERROR(INDEX(Sheet1!$U:$U,MATCH($A{r},Sheet1!$A:$A,0)),"")'.format(r=r))
    if letter == 'W':  return g(r, 'N{r}-'.format(r=r) + BDP.format(r=r, args='"YAS_OAS_SPRD","SETTLE_DT",Company_desc!$A$1'))
    if letter == 'X':  return g(r, 'N{r}-'.format(r=r) + BDP.format(r=r, args='"YAS_OAS_SPRD","SETTLE_DT",Company_desc!$A$2'))
    if letter == 'Y':  return g(r, 'N{r}-'.format(r=r) + BDP.format(r=r, args='"YAS_OAS_SPRD","SETTLE_DT",Company_desc!$D$1'))
    if letter == 'Z':  return g(r, 'K{r}-'.format(r=r) + BDP.format(r=r, args='"YAS_BOND_YLD","YAS_YLD_FLAG=15","SETTLE_DT",Company_desc!$D$1'))
    if letter == 'AA': return g(r, BDP.format(r=r, args='"YAS_OAS_SPRD","SETTLE_DT",Company_desc!$D$1'))
    raise KeyError(letter)

COLS = ['B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U',
        'W','X','Y','Z','AA']

# new headers, styled like the existing X1 header
hdr_font = copy(f['X1'].font); hdr_fill = copy(f['X1'].fill); hdr_align = copy(f['X1'].alignment)
for letter, text in (('Y', 'Spread change since ' + ANCHOR_TXT),
                     ('Z', 'YTW change since ' + ANCHOR_TXT),
                     ('AA', 'Spread on ' + ANCHOR_TXT)):
    c = f[letter + '1']
    c.value = text
    c.font, c.fill, c.alignment = copy(hdr_font), copy(hdr_fill), copy(hdr_align)

body_font = copy(f['W2'].font)
NUMFMT = {'D':'0.000','F':'DD.MM.YYYY','I':'0.###','J':'0.00','K':'0.00','L':'0.00','M':'0.00',
          'N':'0','O':'0.00','P':'0','W':'0','X':'0','Y':'0','Z':'0.00','AA':'0'}
for r in range(2, LAST_ROW + 1):
    for letter in COLS:
        c = f[letter + str(r)]
        c.value = col_formula(letter, r)
        c.font = copy(body_font)
        if letter in NUMFMT:
            c.number_format = NUMFMT[letter]

# extend Table1 to the new last column
t = f.tables['Table1']
t.ref = t.ref.replace('A1:X', 'A1:AA')
t.tableColumns = []
t._initialise_columns()
for i, tc in enumerate(t.tableColumns, start=1):
    v = f.cell(1, i).value
    if v:
        tc.name = str(v)
if t.autoFilter is not None:
    t.autoFilter.ref = t.ref
for letter, w in (('Y', 30), ('Z', 30), ('AA', 24)):
    f.column_dimensions[letter].width = w

# mirror the three new headers on Bonds_list_hard (macro rewrites the body)
for letter in ('Y', 'Z', 'AA'):
    src, dst = f[letter + '1'], blh[letter + '1']
    dst.value = src.value
    dst.font, dst.fill, dst.alignment = copy(src.font), copy(src.fill), copy(src.alignment)
for letter in ('Y', 'Z', 'AA'):
    blh.column_dimensions[letter].width = f.column_dimensions[letter].width

wb.save(OUT)
print('saved', OUT)

# ---------------------------------------------------------------- 3. new output sheet + ReadMe
wb = openpyxl.load_workbook(OUT, keep_vba=True)

if NEW_SHEET in wb.sheetnames:
    del wb[NEW_SHEET]
idx = wb.sheetnames.index('Spread - 1 month') + 1
ws = wb.create_sheet(NEW_SHEET, idx)
ws.sheet_properties.tabColor = 'FF3AB7C8'
ws['A1'] = 'Populated by MacroSPREADSINCE (module SPREAD_SINCE). Run RM_VERSION_WITH_SINCE.'
ws['A1'].font = Font(name='Nunito Sans', sz=12, i=True, color='FF808080')
ws.column_dimensions['A'].width = 70

if 'ReadMe (changes)' in wb.sheetnames:
    del wb['ReadMe (changes)']
rm = wb.create_sheet('ReadMe (changes)')
rm.sheet_properties.tabColor = 'FFFFC000'
rows = [
 ('SYZ Bond Monitor - spread variation since ' + ANCHOR_TXT, ''),
 ('', ''),
 ('WHAT WAS ADDED', ''),
 ('Company_desc!D1', 'Fixed anchor date = ' + ANCHOR_TXT + ' (blue/yellow = the only cell to edit). '
                     'A1 and A2 stay dynamic: TODAY()-7 and TODAY()-30.'),
 ('Formulas!Y', 'Spread change since ' + ANCHOR_TXT + '  =  N - BDP(ISIN,"YAS_OAS_SPRD","SETTLE_DT",Company_desc!$D$1)'),
 ('Formulas!Z', 'YTW change since ' + ANCHOR_TXT + '  =  K - BDP(ISIN,"YAS_BOND_YLD","YAS_YLD_FLAG=15","SETTLE_DT",Company_desc!$D$1)'),
 ('Formulas!AA', 'Spread level on ' + ANCHOR_TXT + '  =  BDP(ISIN,"YAS_OAS_SPRD","SETTLE_DT",Company_desc!$D$1)'),
 ('Formulas!B:U', 'Descriptive columns are now live Bloomberg / lookup formulas instead of being blank '
                  '(see the field list below). Every formula is guarded with IF($A=\"\",\"\",...) so unused '
                  'rows stay empty. Rows 2:769 are pre-filled - paste ISINs in column A only.'),
 ('New sheet', NEW_SHEET + ' - WORST 10 / TOP 10 per currency, ranked on the "since" spread change.'),
 ('New VBA module', 'SPREAD_SINCE.bas - import it (see below). It is purely additive: no existing macro is modified.'),
 ('', ''),
 ('BLOOMBERG FIELDS USED FOR THE DESCRIPTIVE COLUMNS', ''),
 ('C  Company', 'PROPER(BDP(ISIN & " Corp","NAME"))'),
 ('D  Coupon', 'BDP("CPN")'),
 ('E  Structure', 'BDP("PAYMENT_RANK")'),
 ('F  Maturity', 'BDP("MATURITY")'),
 ('G  Country', 'PROPER(BDP("COUNTRY_FULL_NAME"))'),
 ('H  Sector', 'BDP("INDUSTRY_SECTOR")'),
 ('I  Den. (k)', 'BDP("MIN_PIECE")/1000'),
 ('J  Price', 'BDP("PX_LAST")'),
 ('K  YTW', 'BDP("YAS_BOND_YLD","YAS_YLD_FLAG=15")'),
 ('N  Spread', 'BDP("YAS_OAS_SPRD")'),
 ('O  Dur.', 'BDP("YAS_MOD_DUR")'),
 ('R  Currency', 'BDP("CRNCY")'),
 ('Q  Company Description', 'INDEX/MATCH on Company_desc!A:C, falling back to Sheet1!Q (the description library, not a Bloomberg field).'),
 ('B / P / S / U', 'List, Ind LV., Bond Classification, Tax Group are SYZ-internal fields: INDEX/MATCH on Sheet1 by ISIN.'),
 ('T  rating', 'INDEX/MATCH on Sheet1!T, falling back to BDP("RTG_SP") for ISINs not yet in Sheet1.'),
 ('', ''),
 ('HOW TO IMPORT THE MACRO (one-off, ~30 seconds)', ''),
 ('1', 'Open the workbook, press Alt+F11 (VBA editor).'),
 ('2', 'File > Import File... and pick SPREAD_SINCE.bas.'),
 ('3', 'Close the editor and save the workbook as .xlsm.'),
 ('4', 'Run RM_VERSION_WITH_SINCE (Alt+F8) - it refreshes Bonds_list_hard from Formulas (now 27 columns), '
       'then rebuilds SpreadRanking, Spread - 1 week, Spread - 1 month and ' + NEW_SHEET + '.'),
 ('', ''),
 ('NOTES / ASSUMPTIONS', ''),
 ('Header text', 'Formulas!Y1:AA1 spell out ' + ANCHOR_TXT + ' as static text (Excel table headers cannot hold '
                 'formulas). If you change Company_desc!D1, update those three headers too.'),
 ('Column count', 'The original ActualiserDonneesBons copies columns A:X only. '
                  'ActualiserDonneesBonsFull in the new module copies A:AA, so Y/Z/AA reach Bonds_list_hard.'),
 ('Side effect', 'MacroSPREADWEEK / MacroSPREADMONTH already append "column 25 to last column", so the 1-week and '
                 '1-month tables now carry the three "since" columns as extra context. No change was needed there.'),
 ('Recalculation', 'This file was NOT recalculated offline: every value comes from the Bloomberg add-in (_xll.BDP), '
                   'which only resolves on a terminal-connected machine. Open it with the add-in loaded and let Excel calculate.'),
]
rm['A1'].font = Font(name='Nunito Sans', sz=16, b=True)
for i, (a, b) in enumerate(rows, start=1):
    rm.cell(i, 1, a); rm.cell(i, 2, b)
    if a and not b:
        rm.cell(i, 1).font = Font(name='Nunito Sans', sz=12, b=True, color='FF3AB7C8')
    else:
        rm.cell(i, 1).font = Font(name='Nunito Sans', sz=10, b=True)
        rm.cell(i, 2).font = Font(name='Nunito Sans', sz=10)
    rm.cell(i, 2).alignment = Alignment(wrap_text=True, vertical='top')
rm['A1'].font = Font(name='Nunito Sans', sz=16, b=True)
rm.column_dimensions['A'].width = 26
rm.column_dimensions['B'].width = 118

wb.save(OUT)
print('sheets:', wb.sheetnames)

# ---------------------------------------------------------------- 4. embed the VBA source as text
wb = openpyxl.load_workbook(OUT, keep_vba=True)
if 'VBA_SPREAD_SINCE' in wb.sheetnames:
    del wb['VBA_SPREAD_SINCE']
vs = wb.create_sheet('VBA_SPREAD_SINCE')
vs['A1'] = ('Source of SPREAD_SINCE.bas - kept here as a backup. Preferred route: Alt+F11 > '
            'File > Import File... > SPREAD_SINCE.bas. Otherwise Insert > Module and paste rows 3+ below.')
vs['A1'].font = Font(name='Nunito Sans', sz=11, b=True, color='FFC00000')
with open('SPREAD_SINCE.bas', encoding='utf-8') as fh:
    for i, line in enumerate(fh.read().splitlines(), start=3):
        c = vs.cell(i, 1, line)
        c.font = Font(name='Consolas', sz=9)
        c.alignment = Alignment(horizontal='left')
vs.column_dimensions['A'].width = 120
vs.sheet_state = 'hidden'
wb.save(OUT)
print('embedded VBA source rows:', vs.max_row)
